"""Perfilado no destructivo y anonimizado de un libro Excel.

Uso:
    python tools/profile_excel.py <archivo.xlsx> <perfil.json>

El perfil conserva encabezados y catálogos operativos de bajo riesgo, pero
anonimiza muestras de datos y nunca modifica el libro fuente.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
import sys
import zipfile
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries


SENSITIVE_HEADER = re.compile(
    r"(nombre|name|correo|email|mail|tel[eé]fono|phone|celular|m[oó]vil|"
    r"pasaport|document|direcci[oó]n|address|localizador|confirmaci[oó]n|"
    r"cumple|nacimiento|birth|contacto|contact|nota|observaci[oó]n|"
    r"usuario|user|contrase|password|cuenta|credential|secret|token)",
    re.IGNORECASE,
)
CATEGORY_HEADER = re.compile(
    r"(estado|status|canal|channel|moneda|currency|tipo|type|categor[ií]a|"
    r"prioridad|priority|pa[ií]s|country|origen|source|s[ií]/no|yes/no)",
    re.IGNORECASE,
)
EMAIL = re.compile(r"[^\s@]+@[^\s@]+\.[^\s@]+")
PHONEISH = re.compile(r"\+?\d[\d\s().-]{6,}\d")
FORMULA_REF = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_ áéíóúÁÉÍÓÚñÑ.-]+))!")


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return str(value)
    return value


def mask(value: Any) -> str:
    text = str(value)
    digest = hashlib.sha256(text.encode("utf-8", errors="replace")).hexdigest()[:10]
    return f"<anon:{digest};len={len(text)}>"


def safe_header(value: Any, column_index: int) -> str:
    if value is None or str(value).strip() == "":
        return f"__columna_{get_column_letter(column_index)}_sin_encabezado"
    text = re.sub(r"\s+", " ", str(value)).strip()
    if EMAIL.search(text) or PHONEISH.search(text) or len(text) > 120:
        return f"__encabezado_anon_{get_column_letter(column_index)}"
    return text


def value_type(value: Any) -> str:
    if value is None or value == "":
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"
    return type(value).__name__


def formula_text(value: Any) -> str:
    text = getattr(value, "text", None)
    return str(text if text is not None else value)


def normalize_formula(formula: str) -> str:
    result = re.sub(r'"(?:[^"\\]|\\.)*"', '"<texto>"', formula)
    result = re.sub(r"(?<![A-Za-z_])\d+(?:\.\d+)?", "<n>", result)
    return result[:500]


def actual_bounds(ws: Any) -> tuple[int, int, int, int] | None:
    min_row = min_col = None
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)
    if min_row is None or min_col is None:
        return None
    return min_row, min_col, max_row, max_col


def table_headers(ws: Any) -> list[dict[str, Any]]:
    result = []
    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [safe_header(ws.cell(min_row, col).value, col) for col in range(min_col, max_col + 1)]
        result.append(
            {
                "name": table.name,
                "ref": table.ref,
                "header_row": min_row,
                "min_col": min_col,
                "max_col": max_col,
                "headers": headers,
            }
        )
    return result


def header_candidates(ws: Any, bounds: tuple[int, int, int, int] | None) -> list[dict[str, Any]]:
    if not bounds:
        return []
    min_row, min_col, max_row, max_col = bounds
    scan_end = min(max_row, min_row + 49)
    candidates = []
    for row_index in range(min_row, scan_end + 1):
        values = [ws.cell(row_index, col).value for col in range(min_col, max_col + 1)]
        nonempty = [v for v in values if v not in (None, "")]
        strings = [v for v in nonempty if isinstance(v, str) and not str(v).startswith("=")]
        if len(nonempty) < 2 or len(strings) / len(nonempty) < 0.5:
            continue
        unique = len({str(v).strip().casefold() for v in nonempty})
        next_nonempty = 0
        if row_index < max_row:
            next_nonempty = sum(ws.cell(row_index + 1, col).value not in (None, "") for col in range(min_col, max_col + 1))
        score = len(nonempty) * 2 + unique + min(next_nonempty, len(nonempty))
        candidates.append(
            {
                "row": row_index,
                "score": score,
                "nonempty_cells": len(nonempty),
            }
        )
    return sorted(candidates, key=lambda item: (-item["score"], item["row"]))[:3]


def choose_regions(ws: Any, bounds: tuple[int, int, int, int] | None) -> list[dict[str, Any]]:
    tables = table_headers(ws)
    if tables:
        return tables
    candidates = header_candidates(ws, bounds)
    if not candidates or not bounds:
        return []
    min_row, min_col, max_row, max_col = bounds
    best = min(candidates, key=lambda item: item["row"])
    headers = [
        safe_header(ws.cell(best["row"], col).value, col)
        for col in range(min_col, max_col + 1)
    ]
    return [
        {
            "name": "detected_region",
            "ref": f"{get_column_letter(min_col)}{best['row']}:{get_column_letter(max_col)}{max_row}",
            "header_row": best["row"],
            "min_col": min_col,
            "max_col": max_col,
            "headers": headers,
        }
    ]


def profile_column(
    formula_ws: Any,
    value_ws: Any,
    header: str,
    col: int,
    start_row: int,
    end_row: int,
) -> dict[str, Any]:
    values: list[Any] = []
    formulas: list[str] = []
    for row in range(start_row, end_row + 1):
        fcell = formula_ws.cell(row, col)
        if isinstance(fcell, MergedCell):
            continue
        if fcell.data_type == "f" or (isinstance(fcell.value, str) and fcell.value.startswith("=")):
            formulas.append(formula_text(fcell.value))
            values.append(value_ws.cell(row, col).value)
        else:
            values.append(fcell.value)

    types = Counter(value_type(value) for value in values)
    non_null = [value for value in values if value not in (None, "")]
    normalized = [json.dumps(json_value(value), ensure_ascii=False, sort_keys=True) for value in non_null]
    counts = Counter(normalized)
    distinct_count = len(counts)
    duplicate_values = sum(count - 1 for count in counts.values() if count > 1)
    strings = [str(value) for value in non_null if isinstance(value, str)]
    numbers = [float(value) for value in non_null if isinstance(value, (int, float)) and not isinstance(value, bool)]
    dates = [value for value in non_null if isinstance(value, (date, datetime))]
    sensitive = bool(SENSITIVE_HEADER.search(header))
    categorical = bool(CATEGORY_HEADER.search(header)) and distinct_count <= 50 and not sensitive
    if categorical:
        samples = [json.loads(value) for value, _ in counts.most_common(20)]
    else:
        samples = [mask(json.loads(value)) for value, _ in counts.most_common(5)]

    return {
        "column": get_column_letter(col),
        "header": header,
        "sensitive_candidate": sensitive,
        "row_count": len(values),
        "non_null_count": len(non_null),
        "null_count": len(values) - len(non_null),
        "distinct_count": distinct_count,
        "duplicate_occurrences": duplicate_values,
        "type_counts": dict(types),
        "formula_count": len(formulas),
        "formula_patterns": [pattern for pattern, _ in Counter(normalize_formula(f) for f in formulas).most_common(10)],
        "string_length": {
            "min": min((len(value) for value in strings), default=None),
            "max": max((len(value) for value in strings), default=None),
        },
        "numeric_range": {
            "min": min(numbers, default=None),
            "max": max(numbers, default=None),
        },
        "date_range": {
            "min": min(dates).isoformat() if dates else None,
            "max": max(dates).isoformat() if dates else None,
        },
        "samples": [json_value(value) for value in samples],
    }


def profile_sheet(formula_ws: Any, value_ws: Any) -> dict[str, Any]:
    bounds = actual_bounds(formula_ws)
    candidates = header_candidates(formula_ws, bounds)
    regions = choose_regions(formula_ws, bounds)
    formula_counter: Counter[str] = Counter()
    dependencies: Counter[str] = Counter()
    comments = 0
    errors = 0
    total_markers = []
    for row in formula_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.comment is not None:
                comments += 1
            if cell.data_type == "e":
                errors += 1
            if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                formula = formula_text(cell.value)
                formula_counter[normalize_formula(formula)] += 1
                for match in FORMULA_REF.finditer(formula):
                    dependencies[(match.group(1) or match.group(2)).strip()] += 1
            if isinstance(cell.value, str) and re.fullmatch(r"\s*(sub)?total(?:\s+general)?\s*", cell.value, re.I):
                total_markers.append(cell.coordinate)

    region_profiles = []
    for region in regions:
        _, _, _, end_row = range_boundaries(region["ref"])
        columns = []
        for offset, col in enumerate(range(region["min_col"], region["max_col"] + 1)):
            header = region["headers"][offset]
            columns.append(profile_column(formula_ws, value_ws, header, col, region["header_row"] + 1, end_row))
        region_profiles.append({**region, "columns": columns})

    return {
        "title": formula_ws.title,
        "state": formula_ws.sheet_state,
        "declared_dimensions": formula_ws.calculate_dimension(),
        "max_row": formula_ws.max_row,
        "max_column": formula_ws.max_column,
        "actual_bounds": list(bounds) if bounds else None,
        "merged_ranges": [str(item) for item in formula_ws.merged_cells.ranges],
        "hidden_rows": sum(1 for dim in formula_ws.row_dimensions.values() if dim.hidden),
        "hidden_columns": sum(1 for dim in formula_ws.column_dimensions.values() if dim.hidden),
        "freeze_panes": str(formula_ws.freeze_panes) if formula_ws.freeze_panes else None,
        "auto_filter": formula_ws.auto_filter.ref,
        "table_count": len(formula_ws.tables),
        "conditional_format_count": len(formula_ws.conditional_formatting),
        "data_validation_count": len(formula_ws.data_validations.dataValidation),
        "validations": [
            {
                "type": validation.type,
                "operator": validation.operator,
                "sqref": str(validation.sqref),
                "formula1": mask(validation.formula1) if validation.formula1 else None,
                "formula2": mask(validation.formula2) if validation.formula2 else None,
            }
            for validation in formula_ws.data_validations.dataValidation
        ],
        "chart_count": len(formula_ws._charts),
        "image_count": len(formula_ws._images),
        "comment_count": comments,
        "error_cell_count": errors,
        "formula_count": sum(formula_counter.values()),
        "formula_patterns": [
            {"pattern": pattern, "count": count} for pattern, count in formula_counter.most_common(25)
        ],
        "sheet_dependencies": dict(dependencies),
        "total_markers": total_markers[:100],
        "header_candidates": candidates,
        "regions": region_profiles,
    }


def package_inventory(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
    prefixes = {
        "worksheets": "xl/worksheets/",
        "tables": "xl/tables/",
        "charts": "xl/charts/",
        "drawings": "xl/drawings/",
        "pivot_tables": "xl/pivotTables/",
        "pivot_cache": "xl/pivotCache/",
        "external_links": "xl/externalLinks/",
        "connections": "xl/connections",
        "media": "xl/media/",
        "slicers": "xl/slicers/",
    }
    return {
        key: sum(
            name.startswith(prefix)
            and name.endswith(".xml")
            and "/_rels/" not in name
            for name in names
        )
        for key, prefix in prefixes.items()
    } | {
        "macro_project": any(name.endswith("vbaProject.bin") for name in names),
        "custom_xml_parts": sum(
            name.startswith("customXml/")
            and name.endswith(".xml")
            and "/_rels/" not in name
            for name in names
        ),
    }


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Uso: python tools/profile_excel.py <archivo.xlsx> <perfil.json>")
    source = Path(sys.argv[1]).resolve()
    destination = Path(sys.argv[2]).resolve()
    source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    formula_book = openpyxl.load_workbook(source, data_only=False, read_only=False, keep_links=True)
    value_book = openpyxl.load_workbook(source, data_only=True, read_only=False, keep_links=True)

    profile = {
        "source": {
            "name": source.name,
            "size_bytes": source.stat().st_size,
            "sha256": source_hash,
            "modified_utc": datetime.utcfromtimestamp(source.stat().st_mtime).isoformat() + "Z",
        },
        "workbook": {
            "sheet_count": len(formula_book.sheetnames),
            "sheet_order": formula_book.sheetnames,
            "epoch": str(formula_book.epoch),
            "defined_names": [
                {"name": name, "type": type(item).__name__}
                for name, item in formula_book.defined_names.items()
            ],
            "package_parts": package_inventory(source),
        },
        "sheets": [
            profile_sheet(formula_book[name], value_book[name]) for name in formula_book.sheetnames
        ],
    }
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "profile": str(destination),
                "sha256": source_hash,
                "sheets": len(formula_book.sheetnames),
                "sheet_names": formula_book.sheetnames,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
