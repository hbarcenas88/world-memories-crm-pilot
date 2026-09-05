import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table

from tools.profile_excel import (
    actual_bounds,
    choose_regions,
    formula_text,
    header_candidates,
    package_inventory,
    profile_column,
    table_headers,
)


class TableHeaderTests(unittest.TestCase):
    def test_reads_table_object_instead_of_table_range_string(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["ID", "Estado"])
        sheet.append([1, "Nuevo"])
        sheet.add_table(Table(displayName="Leads", ref="A1:B2"))

        result = table_headers(sheet)

        self.assertEqual(result[0]["name"], "Leads")
        self.assertEqual(result[0]["ref"], "A1:B2")
        self.assertEqual(result[0]["headers"], ["ID", "Estado"])


class HeaderDetectionPrivacyTests(unittest.TestCase):
    def test_does_not_emit_values_from_candidate_data_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nombre", "Estado"])
        sheet.append(["Persona Sensible", "Nuevo", "Dato adicional"])

        candidates = header_candidates(sheet, actual_bounds(sheet))

        self.assertNotIn("Persona Sensible", str(candidates))

    def test_uses_earliest_structural_row_when_no_excel_table_exists(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nombre", "Estado"])
        sheet.append(["Persona Sensible", "Nuevo", "Dato adicional"])

        regions = choose_regions(sheet, actual_bounds(sheet))

        self.assertEqual(regions[0]["header_row"], 1)
        self.assertEqual(regions[0]["headers"][:2], ["Nombre", "Estado"])

    def test_marks_credential_columns_as_sensitive(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "secreto-de-prueba"

        result = profile_column(sheet, sheet, "Contraseña", 1, 1, 1)

        self.assertTrue(result["sensitive_candidate"])


class FormulaExtractionTests(unittest.TestCase):
    def test_extracts_text_from_array_formula(self):
        formula = ArrayFormula(ref="B2", text="=XLOOKUP(A2,Leads!A:A,Leads!B:B)")

        self.assertEqual(formula_text(formula), "=XLOOKUP(A2,Leads!A:A,Leads!B:B)")


class PackageInventoryTests(unittest.TestCase):
    def test_does_not_count_relationship_files_as_worksheets(self):
        Path("artifacts").mkdir(exist_ok=True)
        package = Path("artifacts/package_inventory_test.xlsx")
        try:
            with zipfile.ZipFile(package, "w") as archive:
                archive.writestr("xl/worksheets/sheet1.xml", "<worksheet />")
                archive.writestr("xl/worksheets/_rels/sheet1.xml.rels", "<Relationships />")

            result = package_inventory(package)
        finally:
            package.unlink(missing_ok=True)

        self.assertEqual(result["worksheets"], 1)


if __name__ == "__main__":
    unittest.main()
